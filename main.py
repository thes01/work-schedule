"""Monthly nurse scheduling using CP-SAT.

Key rules (updated):
* Shifts: Day (D, 11h; nurse 1 Mon-Thu 10h), Night (N, 12h), R8 (8h morning).
* Horizon: 31 days (Mon start): 4 full weeks + 3 extra workdays.
* Base nurses: 19 (indices 0..18). Extra 8h-only nurses: 1 (index 19). Nurse numbering displayed 1-based.
* Workday demand: total day-like (D+R8) in [9,10]; exactly 1 night.
* Weekend demand: day D in [5,6]; exactly 1 night; extra nurses off.
* Base nurse hours normally in [143,146] except:
		- Nurse 1 (index 0): custom 10h for Mon-Thu day shifts.
        - Nurse 3 (index 2): NO night shifts allowed (pure day-like).
* Extra nurse (nurse 20): exactly 136h (17×8h); R8 only.
* Night shift limits: each base nurse (except nurse 3) must have at least 1 and at most 3 night shifts; nurse 3 none; rest rules for singletons and pairs (N,N) with required off days.
* No 3 consecutive day-like shifts (D) for any base nurse.
* Weekend workload balanced: each base nurse has 2 or 3 weekend shifts; pairwise difference ≤ 1.
* Objective: minimize (1) sum of absolute deviations from target hours, (2) night distribution deviation, (3) surplus day-like shifts above theoretical minimum.

Outputs: textual summary + Excel schedule with color coding and legends.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import List, Dict, Tuple
import csv
from pathlib import Path

try:
	from openpyxl import Workbook
	from openpyxl.styles import PatternFill, Alignment, Font
	from openpyxl.comments import Comment
	from openpyxl.utils import get_column_letter
except ImportError:  # graceful fallback if not installed yet
	Workbook = None  # type: ignore
from ortools.sat.python import cp_model


@dataclass(frozen=True)
class ProblemData:
	# Original (base) nurses who can perform Day (D) and Night (N) shifts
	base_nurses: int = 21
	# Additional nurses performing only 8h shifts (R8) on workdays; no nights, no weekends
	extra_nurses: int = 1
	num_days: int = 31  # 4 weeks (28) + 3 workdays
	day_shift_hours: int = 11  # Standard day hours (nurse 1 has custom 10h Mon-Thu non-weekend)
	night_shift_hours: int = 11 
	eight_hour_shift_hours: int = 8  # R8
	min_hours: int = 140  # base nurse min hours
	max_hours: int = 146  # base nurse max hours
	target_hours: int = 145  # balancing target (base nurses only)
	workday_shift_min: int = 9
	workday_shift_max: int = 10
	weekend_shift_min: int = 6
	weekend_shift_max: int = 6
    
	@property
	def num_nurses(self) -> int:
		return self.base_nurses + self.extra_nurses


def build_and_solve(data: ProblemData) -> None:
	"""Builds and solves the extended model with extra 8h shift nurses (R8)."""
	model = cp_model.CpModel()

	# Indices
	base_nurses = list(range(data.base_nurses))
	extra_nurses = list(range(data.base_nurses, data.base_nurses + data.extra_nurses))
	nurses = base_nurses + extra_nurses
	days = list(range(data.num_days))

	# Shift indices
	D, N, R8 = 0, 1, 2

	# Helper: weekend detection
	def is_weekend(day: int) -> bool:
		if day >= 28:
			return False
		dow = day % 7
		return dow in (5, 6)

	workdays: List[int] = [d for d in days if not is_weekend(d)]
	weekend_days: List[int] = [d for d in days if is_weekend(d)]

	# Allowed shifts per nurse
	nurse_shifts: Dict[int, Tuple[int, ...]] = {}
	for n in nurses:
		if n in base_nurses:
			# Base nurses can take Day or Night shifts
			# Nurse index 2 (number 3) is forbidden from night shifts.
			if n == 2:
				nurse_shifts[n] = (D,)
			else:
				nurse_shifts[n] = (D, N)
		else:
			# Extra nurse (index 19, nurse number 20) is R8-only
			nurse_shifts[n] = (R8,)

	# Decision variables only for allowed shifts
	x: Dict[Tuple[int, int, int], cp_model.IntVar] = {}
	for n in nurses:
		for d in days:
			for s in nurse_shifts[n]:
				label = {D: 'D', N: 'N', R8: 'R8'}[s]
				x[(n, d, s)] = model.new_bool_var(f"x_n{n}_d{d}_{label}")

	# 1) At most one shift per nurse per day
	for n in nurses:
		for d in days:
			model.add_at_most_one(x[(n, d, s)] for s in nurse_shifts[n])

	# 2) Daily demand constraints
	# Workdays: (D + R8) in [workday_shift_min, workday_shift_max]; exactly 1 night (base nurses only).
	for d in workdays:
		day_like = []
		day_like.extend(x[(n, d, D)] for n in base_nurses if (n, d, D) in x)
		# extra nurse R8
		day_like.extend(x[(n, d, R8)] for n in extra_nurses if (n, d, R8) in x)
		model.add_linear_constraint(sum(day_like), data.workday_shift_min, data.workday_shift_max)
		model.add(sum(x[(n, d, N)] for n in base_nurses if (n, d, N) in x) == 1)
	# Weekends: (D) in [5,6]; exactly 1 night; extra nurses off.
	for d in weekend_days:
		model.add_linear_constraint(sum(x[(n, d, D)] for n in base_nurses if (n, d, D) in x), data.weekend_shift_min, data.weekend_shift_max)
		model.add(sum(x[(n, d, N)] for n in base_nurses if (n, d, N) in x) == 1)
		for n in extra_nurses:
			for s in nurse_shifts[n]:
				model.add(x[(n, d, s)] == 0)

	# 3) Hours per nurse
	hours: List[cp_model.LinearExpr] = [0] * len(nurses)
	deviations: List[cp_model.IntVar] = []  # base nurses only
	for n in nurses:
		if n in base_nurses:
			# Day hours with custom rule for nurse index 0 (Mon-Thu =10h)
			day_terms = []
			for d in days:
				if (n, d, D) in x:
					coef = data.day_shift_hours
					if n == 0:
						dow = d % 7
						if (not is_weekend(d)) and dow in (0, 1, 2, 3):
							coef = 10
					day_terms.append(coef * x[(n, d, D)])
			day_hours = sum(day_terms) if day_terms else 0
			night_hours = data.night_shift_hours * sum(x[(n, d, N)] for d in days if (n, d, N) in x)
			h = day_hours + night_hours
			hours[n] = h
			# All base nurses have same hour constraints
			model.add_linear_constraint(h, data.min_hours, data.max_hours)
			pos = model.new_int_var(0, data.max_hours - data.target_hours, f"dev_pos_{n}")
			neg = model.new_int_var(0, data.target_hours - data.min_hours, f"dev_neg_{n}")
			model.add(h - data.target_hours == pos - neg)
			dev = model.new_int_var(0, data.max_hours - data.min_hours, f"dev_abs_{n}")
			model.add(dev == pos + neg)
			deviations.append(dev)
		else:
			# Extra nurse exact 17 * 8h = 136h
			shift_sum = sum(x[(n, d, s)] for d in workdays for s in nurse_shifts[n] if (n, d, s) in x)
			h = data.eight_hour_shift_hours * shift_sum
			hours[n] = h
			model.add(h == 17 * data.eight_hour_shift_hours)

	# 4) No 3 consecutive day-like (D) shifts for the base nurses
	for n in base_nurses:
		for d in range(data.num_days - 2):
			# Variables (n, d+i, D) always exist for base nurses due to nurse_shifts definition
			model.add(
				x[(n, d, D)]
				+ x[(n, d + 1, D)]
				+ x[(n, d + 2, D)]
				<= 2
			)

	# 5) Max 3 night shifts per base nurse
	for n in base_nurses:
		if n == 2:
			# No night shifts allowed; enforce zero if any N var accidentally created.
			for d in days:
				if (n, d, N) in x:
					model.add(x[(n, d, N)] == 0)
			continue
		# Each other base nurse: at least 1, at most 3 nights
		night_sum = sum(x[(n, d, N)] for d in days if (n, d, N) in x)
		model.add(night_sum >= 1)
		model.add(night_sum <= 3)

	# 6) Night shift rest rules (base nurses only)
	for n in base_nurses:
		pair_vars = []
		for d in range(data.num_days - 1):
			if (n, d, N) in x and (n, d + 1, N) in x:
				p = model.new_bool_var(f"pair_n{n}_d{d}")
				pair_vars.append(p)
				model.add(p <= x[(n, d, N)])
				model.add(p <= x[(n, d + 1, N)])
				model.add(x[(n, d, N)] + x[(n, d + 1, N)] - 1 <= p)
			else:
				pair_vars.append(model.new_constant(0))  # filler
		for d in range(data.num_days - 2):
			model.add(pair_vars[d] + pair_vars[d + 1] <= 1)
		singleton = []
		for d in days:
			if (n, d, N) in x:
				s = model.new_bool_var(f"single_n{n}_d{d}")
				singleton.append(s)
				model.add(s <= x[(n, d, N)])
				if d - 1 >= 0:
					model.add(s + pair_vars[d - 1] <= 1)
				if d < data.num_days - 1:
					model.add(s + pair_vars[d] <= 1)
				lin_expr = x[(n, d, N)]
				if d - 1 >= 0:
					lin_expr -= pair_vars[d - 1]
				if d < data.num_days - 1:
					lin_expr -= pair_vars[d]
				model.add(lin_expr - s <= 0)
			else:
				singleton.append(model.new_constant(0))
		# Pair rest
		for d in range(data.num_days - 1):
			p = pair_vars[d]
			if isinstance(p, cp_model.IntVar):
				if d - 1 >= 0:
					if (n, d - 1, D) in x:
						model.add(x[(n, d - 1, D)] + p <= 1)
				for k in (2, 3):
					if d + k < data.num_days:
						if (n, d + k, D) in x:
							model.add(x[(n, d + k, D)] + p <= 1)
						if (n, d + k, N) in x:
							model.add(x[(n, d + k, N)] + p <= 1)
		# Singleton rest (two days off after)
		for d in days:
			s = singleton[d]
			if isinstance(s, cp_model.IntVar):
				for k in (1, 2):
					if d + k < data.num_days:
						if (n, d + k, D) in x:
							model.add(x[(n, d + k, D)] + s <= 1)
						if (n, d + k, N) in x:
							model.add(x[(n, d + k, N)] + s <= 1)

	# 7) Weekend balancing (base nurses only)
	weekend_shifts = []
	for n in base_nurses:
		w = sum(x[(n, d, s)] for d in weekend_days for s in (D, N) if (n, d, s) in x)
		weekend_shifts.append(w)
		model.add_linear_constraint(w, 2, 3)
	for i in range(len(base_nurses)):
		for j in range(i + 1, len(base_nurses)):
			model.add(weekend_shifts[i] - weekend_shifts[j] <= 1)
			model.add(weekend_shifts[j] - weekend_shifts[i] <= 1)

	# 8) Objective components
	total_day_shifts = (
		sum(x[(n, d, D)] for n in base_nurses for d in days if (n, d, D) in x)
		+ sum(x[(n, d, R8)] for n in extra_nurses for d in workdays if (n, d, R8) in x)
	)
	min_needed_day_shifts = 9 * len(workdays) + 5 * len(weekend_days)
	surplus = model.new_int_var(0, data.num_days * data.num_nurses, "day_surplus")
	model.add(total_day_shifts - min_needed_day_shifts == surplus)

	avg_night_times100 = int(100 * (len(days) / max(1, data.base_nurses)))  # informational only
	night_devs = []
	for n in base_nurses:
		if n == 2:  # skip nurse 3 (no nights); treat as zero deviation implicitly
			continue
		nc = sum(x[(n, d, N)] for d in days if (n, d, N) in x)
		pos = model.new_int_var(0, 3, f"night_pos_{n}")
		neg = model.new_int_var(0, 2, f"night_neg_{n}")
		model.add(nc - 2 == pos - neg)
		nd = model.new_int_var(0, 3, f"night_dev_{n}")
		model.add(nd == pos + neg)
		night_devs.append(nd)

	# Weights
	W_HOURS = 100
	W_NIGHT_DEV = 10
	W_SURPLUS = 1
	objective_terms = [W_HOURS * sum(deviations), W_NIGHT_DEV * sum(night_devs), W_SURPLUS * surplus]
	model.minimize(sum(objective_terms))

	# Solve
	solver = cp_model.CpSolver()
	solver.parameters.max_time_in_seconds = 60.0
	solver.parameters.num_search_workers = 8
	status = solver.solve(model)
	status_name = solver.status_name(status)
	print(f"Solver status: {status_name}")
	if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
		print("No feasible solution found under current constraints.")
		return

	print("\n=== Summary ===")
	total_hours = 0
	for n in nurses:
		h_val = solver.value(hours[n])
		total_hours += h_val
		day_like_count = 0
		night_count = 0
		for d in days:
			for s in nurse_shifts[n]:
				if s in (D, R8) and (n, d, s) in x and solver.boolean_value(x[(n, d, s)]):
					day_like_count += 1
				if s == N and (n, d, N) in x and solver.boolean_value(x[(n, d, N)]):
					night_count += 1
		wkd = sum(
			solver.value(x[(n, d, s)])
			for d in weekend_days
			for s in nurse_shifts[n]
			if (n, d, s) in x
		)
		if n == 2:
			label = f"Nurse {n:2d} (no nights):"
		else:
			label = f"Nurse {n:2d}:"
		print(
			f"{label} hours={h_val} day_like_shifts={day_like_count} night_shifts={night_count} weekend_shifts={wkd}"
		)
	print(f"Total hours (all nurses): {total_hours}")
	print(f"Total day-like shifts: {solver.value(total_day_shifts)} (min theoretical {min_needed_day_shifts})")
	print(f"Objective value: {solver.objective_value}")

	# Excel export (extended with distinct R8 formatting)
	if Workbook is None:
		print("openpyxl not installed: skipping Excel export. Install 'openpyxl' to enable.")
	else:
		wb = Workbook()
		ws = wb.active
		ws.title = "Schedule"
		ws.cell(row=1, column=1, value="Pečovatelka")
		for d in days:
			ws.cell(row=1, column=2 + d, value=d + 1)

		day_counts = [0] * data.num_days  # total day-like (D + R8)
		night_counts = [0] * data.num_days

		WE_FILL = PatternFill(start_color="FFE9CC", end_color="FFE9CC", fill_type="solid")
		D_FILL = PatternFill(start_color="E0F4FF", end_color="E0F4FF", fill_type="solid")
		N_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
		R8_FILL = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")  # light green
		HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

		for c in range(1, data.num_days + 2):
			cell = ws.cell(row=1, column=c)
			cell.fill = HEADER_FILL
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center")

		for n in nurses:
			row_index = 2 + n
			ws.cell(row=row_index, column=1, value=n + 1)
			row_day_like = 0
			row_night = 0
			row_hours = 0
			for d in days:
				cell = ws.cell(row=row_index, column=2 + d)
				val_display = ""
				is_we = is_weekend(d)
				if (n, d, N) in x and solver.boolean_value(x[(n, d, N)]):
					val_display = str(data.night_shift_hours)
					row_night += 1
					row_hours += data.night_shift_hours
					night_counts[d] += 1
					cell.fill = N_FILL if not is_we else N_FILL  # extra nurses don't work weekends anyway
				elif (n, d, D) in x and solver.boolean_value(x[(n, d, D)]):
					# Day shift hours (may be 10 for nurse 0 Mon-Thu)
					if n == 0:
						dow = d % 7
						if (not is_we) and dow in (0,1,2,3):
							val_display = "10"
							row_hours += 10
						else:
							val_display = str(data.day_shift_hours)
							row_hours += data.day_shift_hours
					else:
						val_display = str(data.day_shift_hours)
						row_hours += data.day_shift_hours
					row_day_like += 1
					day_counts[d] += 1
					cell.fill = D_FILL
				elif (n, d, R8) in x and solver.boolean_value(x[(n, d, R8)]):
					val_display = "R8"
					row_day_like += 1
					row_hours += data.eight_hour_shift_hours
					day_counts[d] += 1
					cell.fill = R8_FILL
				else:
					if is_we:
						cell.fill = WE_FILL
				if val_display:
					cell.value = val_display
				cell.alignment = Alignment(horizontal="center")
			# Summaries per row
			ws.cell(row=row_index, column=2 + data.num_days, value=row_hours)
			ws.cell(row=row_index, column=3 + data.num_days, value=row_day_like + row_night)

		# Comments / Legend
		first_nurse_cell = ws.cell(row=2, column=1)
		first_nurse_cell.comment = Comment("Pečovatelka 1: Po-Čt denní směny 10h; Pátek/víkend 11h; noční 12h", "System")
		if extra_nurses:
			first_extra = extra_nurses[0]  # nurse 20 (index 19) R8-only
			cell_extra = ws.cell(row=2 + first_extra, column=1)
			cell_extra.comment = Comment("Pečovatelka 20: Pouze R8.", "System")
		# Comment for nurse 3 no nights
		no_night_cell = ws.cell(row=2 + 2, column=1)
		no_night_cell.comment = Comment("Pečovatelka 3: Bez nočních směn.", "System")

		# Footer rows (aggregated counts)
		day_sum_row = 2 + data.num_nurses + 1
		night_sum_row = day_sum_row + 1
		label_day = ws.cell(row=day_sum_row, column=1, value="Denní směny")
		label_night = ws.cell(row=night_sum_row, column=1, value="Noční směny")
		label_day.font = Font(bold=True)
		label_night.font = Font(bold=True)
		for d in days:
			ws.cell(row=day_sum_row, column=2 + d, value=day_counts[d])
			ws.cell(row=night_sum_row, column=2 + d, value=night_counts[d])
		for c in range(1, 2 + data.num_days):
			ws.cell(row=day_sum_row, column=c).alignment = Alignment(horizontal="center")
			ws.cell(row=night_sum_row, column=c).alignment = Alignment(horizontal="center")

		ws.cell(row=1, column=2 + data.num_days, value="Hodin").font = Font(bold=True)
		ws.cell(row=1, column=3 + data.num_days, value="Směn celkem").font = Font(bold=True)
		ws.cell(row=day_sum_row, column=2 + data.num_days, value="-")
		ws.cell(row=night_sum_row, column=2 + data.num_days, value="-")
		ws.cell(row=day_sum_row, column=3 + data.num_days, value="-")
		ws.cell(row=night_sum_row, column=3 + data.num_days, value="-")

		# Legend (placed below summaries)
		legend_row = night_sum_row + 2
		ws.cell(row=legend_row, column=1, value="Legenda:")
		ws.cell(row=legend_row, column=7, value="R8=Ranní 8h")
		ws.cell(row=legend_row, column=7).fill = R8_FILL

		from openpyxl.utils import get_column_letter
		for col in range(1, 4 + data.num_days):
			col_letter = get_column_letter(col)
			ws.column_dimensions[col_letter].width = 5 if col > 1 else 10
		excel_path = Path("schedule_extended.xlsx")
		wb.save(excel_path)
		print(f"\nExtended Excel exported to {excel_path}")

	print("\n=== Solver Stats ===")
	print(f"Conflicts: {solver.num_conflicts}")
	print(f"Branches : {solver.num_branches}")
	print(f"Wall time: {solver.wall_time:.2f}s")
	# avg_night_times100 kept for potential debugging output


def main() -> None:  # Entry point
	data = ProblemData()
	build_and_solve(data)


if __name__ == "__main__":  # pragma: no cover
	main()
